#! python3
# blankRowInserter.py - Takes two integers and inserts blank rows
# 3 arguments - 1st integer for which row to insert, 2nd for how many blank lines
# 3rd argument is the name of the xlsx file

import sys, openpyxl, os
from openpyxl.utils import get_column_letter as N2L

os.chdir('C:\\users\\daniel.castillo\\desktop\\python learning')

while True:
    if len(sys.argv) == 4:
        try:
            rowToInsert = int(sys.argv[2])
            blankLines = int(sys.argv[3])
            filePath = str(sys.argv[4])
            break
        except Exception as err:
            print(str(err))
            print('You dumbass. It is two numbers and then the file path')
            
    else:
        try:
            print('What is the file path? Be absolute: ')
            filePath = input()
            print('At which row do you want to add blank lines?')
            rowToInsert = int(input())
            print('How many blank lines do you want to insert?')
            blankLines = int(input())
            break
        except Exception as err:
            print(err)
            print('Dumbass, use numbers for the last two questions.')


#print(str(rowToInsert) + ' ' + str(blankLines) + ' ' + filePath)

wb = openpyxl.load_workbook(filePath)
sheet = wb.active
firstSlice = sheet['A1':str(N2L(sheet.max_column)) + str(rowToInsert - 1)]
lastSlice = sheet['A' + str(rowToInsert):str(N2L(sheet.max_column)) + str(sheet.max_row)]
wb2 = openpyxl.Workbook()
sheet2 = wb2.active



for row in firstSlice:
    for cell in row:
        sheet2[cell.coordinate] =  cell.value
        
for row in lastSlice:
    for cell in row:
        #print(N2L(cell.column) + str(cell.row + blankLines))
        sheet2[N2L(cell.column) + str(cell.row + blankLines)] = cell.value    

wb2.save('New file with %s blank lines.xlsx' % (blankLines))
print('Done.')
