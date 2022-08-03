#! python3
# spreadsheetInverter.py - Inverts the X and Y of a spreadsheet

import sys, openpyxl, os
from openpyxl.utils import get_column_letter as N2L

os.chdir('C:\\users\\daniel.castillo\\desktop\\python learning')

if len(sys.argv) == 2:
    filePath = sys.argv[2]
else:
    print('Path of the spreadsheet you want to invert: ' )
    filePath = input()

wb = openpyxl.load_workbook(filePath)
sheet = wb.active
wholeSlice = sheet['A1':str(N2L(sheet.max_column)) + str(sheet.max_row)]
wb2 = openpyxl.Workbook()
sheet2 = wb2.active

for row in wholeSlice:
    for cell in row:
        sheet2[str(N2L(cell.row)) + str(cell.column)] = cell.value

wb2.save('New file inverted.xlsx')
print('done.')
