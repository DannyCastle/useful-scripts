#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl, os

os.chdir('C:\\users\\daniel.castillo\\desktop\\python learning')

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce and their updated prices

PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

#Loop through the rows and update prices

for rowNum in range(2, sheet.max_row): #skips first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
