#! python3
# spreadsheet2Text.py - Inverse of other program. Moves column values to text files

import sys, openpyxl, os
from openpyxl.utils import get_column_letter as N2L

os.chdir('C:\\users\\daniel.castillo\\desktop\\python learning')

if len(sys.argv) == 2:
    excelPath = sys.argv[2]
else:
    print('What is the path of the Excel file?')
    excelPath = input()
    
wb = openpyxl.load_workbook(excelPath)
sheet = wb.active
colCounter = 1

for col in range(1, sheet.max_column + 1):
    file = open('Column %s rows.txt' % (N2L(col)), 'w')
    for row in range(1, sheet.max_row + 1):
        file.write(sheet[N2L(col) + str(row)].value)
    file.close()
