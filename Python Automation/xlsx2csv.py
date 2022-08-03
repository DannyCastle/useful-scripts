#! python3
# xlsx2csv.py - Converts all xlsx files in a folder to csv


import openpyxl, os, csv
from openpyxl.utils import get_column_letter as N2L

os.chdir('C:\\users\\daniel.castillo\\desktop\\python learning\\xlsx2csv')

os.makedirs('new CSVs', exist_ok=True)

for excelFile in os.listdir('.'):
    if not excelFile.endswith('.xlsx'):
        continue

    print('Converting %s to .csv......' % (excelFile))
    wb = openpyxl.load_workbook(excelFile)
    for sheet in wb:
        csvFileName = os.path.splitext(excelFile)[0] + '_' + sheet.title + '.csv'
        csvFile = open(os.path.join('new CSVs', csvFileName), 'w', newline='')
        csvWriter = csv.writer(csvFile)
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []
            for colNum in range(1, sheet.max_column + 1):
                rowData.append(sheet[N2L(colNum) + str(rowNum)].value)
            csvWriter.writerow(rowData)
        csvFile.close()
